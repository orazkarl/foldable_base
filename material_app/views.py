from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import generic
from construction_objects_app.models import ConstructionObject, InvoiceForPayment, Contract
from .models import Material, ReleasedMaterial, ReleasedMaterialItem
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from openpyxl import load_workbook
import requests
from django.db.models import F, Q
from transliterate import slugify
import telebot
import datetime

import os
from foldable_base.settings import BASE_DIR
from docxtpl import DocxTemplate

GENERAL_BOT_TOKEN = '1270115367:AAGCRLBP1iSZhpTniwVYQ9p9fqLysY668ew'
channel_id = '-1001342160485'


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AddMaterialView(generic.TemplateView):
    template_name = 'construction_objects/contract/request/invoice/add_material.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        invoice = InvoiceForPayment.objects.get(id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'manager' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        self.extra_context = {
            'construction_object': construction_object,
            'invoice': invoice,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        invoice = InvoiceForPayment.objects.get(id=int(request.POST['id']))
        object_name = invoice.request_for_material.contract.construction_object.name
        object_name = object_name.split(' ')
        object_slug = ''

        for i in object_name:
            if slugify(i) != None:
                i = slugify(i)
            object_slug += i[0]
        object_slug = object_slug.upper()
        doc_file = request.FILES['doc_file']
        wb = load_workbook(doc_file)
        sheet_name = wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        data = list(sheet.values)
        for item in data:
            name = item[0]
            quantuty = item[1]
            units = item[2]
            price = item[3]
            sum_price = item[4]
            instriment_code = None
            if len(item) > 5:
                instriment_code = item[5]

            material = Material.objects.create(invoice=invoice, name=name, quantity=quantuty, units=units, price=price,
                                               sum_price=sum_price)

            if instriment_code == 1:
                instriment_code = 'I' + object_slug + '-' + str(datetime.datetime.now().year) + '-'
                material.instrument_code = instriment_code + str(material.id)
                material.is_instrument = True
            material.save()

        return redirect('/request_for_material/detail/' + str(invoice.request_for_material.id))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class PaidMaterailsView(generic.TemplateView):
    template_name = 'materials/paid_materials/invoices.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        invoices = InvoiceForPayment.objects.filter(
            request_for_material__contract__construction_object=construction_object)

        self.extra_context = {
            'construction_object': construction_object,
            'invoices': invoices,
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InvoicePaidMaterialsView(generic.TemplateView):
    template_name = 'materials/paid_materials/paid_materials.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        materials = Material.objects.filter(invoice__id=int(self.kwargs['id'])).filter(~Q(ok=F('quantity')))
        invoice = InvoiceForPayment.objects.get(id=self.kwargs['id'])

        if list(materials) == [] and list(invoice.material.all()) != []:
            invoice.is_done = True
            invoice.save()

        self.extra_context = {
            'construction_object': construction_object,
            'materials': materials,
            'invoice': invoice
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        materials = request.POST.getlist('materials')
        materials = Material.objects.filter(id__in=materials)
        context = {
            'materials': materials,
            'construction_object': construction_object,
        }
        if request.POST['submit'] == 'delivered':
            for material in materials:
                material.is_delivery = True
                material.status = 'ок'
                material.ok = material.quantity
                material.marriage, material.shortage, material.inconsistency = 0, 0, 0
                material.save()
        elif request.POST['submit'] == 'marriage':
            return render(request, template_name='materials/paid_materials/marriage_materials.html',
                          context=context)

        elif request.POST['submit'] == 'return':
            return render(request, template_name='materials/paid_materials/return_materials.html',
                          context=context)
        return redirect('/construction_objects/invoice/' + str(self.kwargs['id']) + '/materials')


def marriage_materials(request):
    if request.POST:
        message = '🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔\n'
        message += 'Бракованные товары\n'
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            material.ok = request.POST['ok' + str(i)]
            material.marriage = request.POST['brak' + str(i)]
            material.shortage = request.POST['neso' + str(i)]
            material.inconsistency = request.POST['nexv' + str(i)]
            material.status = 'брак'
            if int(material.marriage) + int(material.ok) + int(material.shortage) + int(
                    material.inconsistency) != material.quantity:
                return HttpResponse('Ошибка!')
            if int(material.ok) > 1:
                material.is_delivery = True
            material.save()

            message += str(
                i) + '. ' + material.name + '\n' + 'Ок: ' + material.ok + '\nБрак: ' + material.marriage + '\nНехватка: ' + material.inconsistency + '\nНесоответсвие: ' + material.shortage + '\n\n'
            invoice = material.invoice
        if request.POST['comment']:
            message += 'Примечение: ' + request.POST['comment'] + '\n'
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment=invoice)
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        message += 'БИН: ' + invoice.bin + '\n'
        message += 'Название компании: ' + invoice.name_company + '\n'
        requests.get("https://api.telegram.org/bot%s/sendMessage" % GENERAL_BOT_TOKEN,
                     params={'chat_id': '-1001342160485', 'text': message})
        bot = telebot.TeleBot(GENERAL_BOT_TOKEN)
        bot.send_document(channel_id, invoice.doc_file)
    return redirect(request.META.get('HTTP_REFERER'))


def return_materials(request):
    if request.POST:
        message = '🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔\n'
        message += 'Возврат товаров\n'
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            return_materials_count = int(request.POST['return' + str(i)])
            invoice = material.invoice
            if return_materials_count > material.quantity:
                return HttpResponse('Ошибка!')
            elif return_materials_count == material.quantity:
                material.delete()
            else:
                material.quantity = material.quantity - return_materials_count
                material.marriage, material.shortage, material.inconsistency = 0, 0, 0
                material.ok = material.quantity
                material.status = 'ок'
                material.save()
            message += str(i) + '. ' + material.name + '\n' + 'Количество: ' + str(
                material.quantity) + '\nКоличество возвратных товаров: ' + str(return_materials_count) + '\n\n'
        if request.POST['comment']:
            message += 'Примечение: ' + request.POST['comment']
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment=invoice)
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        message += 'БИН: ' + str(invoice.bin) + '\n'
        message += 'Название компании: ' + invoice.name_company + '\n'
        requests.get("https://api.telegram.org/bot%s/sendMessage" % GENERAL_BOT_TOKEN,
                     params={'chat_id': '-1001342160485', 'text': message})
        bot = telebot.TeleBot(GENERAL_BOT_TOKEN)
        bot.send_document(channel_id, invoice.doc_file)
    return redirect(request.META.get('HTTP_REFERER'))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ContractMaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/contracts.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        contracts = Contract.objects.filter(construction_object=construction_object)
        self.extra_context = {
            'construction_object': construction_object,
            'contracts': contracts
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class MaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/materials.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(contract__slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        materials = Material.objects.filter(invoice__request_for_material__contract__slug=self.kwargs['slug'],
                                            is_delivery=True, invoice__is_done=True, instrument_code=None)

        self.extra_context = {
            'construction_object': construction_object,
            'materials': materials,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        materials = request.POST.getlist('materials')
        materials = Material.objects.filter(id__in=materials)
        construction_object = ConstructionObject.objects.get(contract__slug=self.kwargs['slug'])
        context = {
            'materials': materials,
            'construction_object': construction_object,
        }
        return render(request, template_name='materials/store_mateials/release_materials.html',context=context)


def release_materials(request):
    if request.POST:
        contract = Material.objects.get(id=int(request.POST['material1'])).invoice.request_for_material.contract
        construction_object = ConstructionObject.objects.get(contract=contract)
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        is_instrument = False
        released_material = ReleasedMaterial.objects.create(user=request.user, release_date=datetime.datetime.now(),contract=contract)
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            released_materials_count = int(request.POST['release' + str(i)])
            material.remainder_count = material.quantity - material.release_count
            material.release_count = material.release_count + released_materials_count
            is_instrument = material.is_instrument
            material.save()
            ReleasedMaterialItem.objects.create(released_material=released_material, material=material,release_count=released_materials_count)
        unique_code = ''
        for i in construction_object.name.split(' '):
            if slugify(i) != None:
                i = slugify(i)
            unique_code += i[0]
        unique_code = unique_code.upper()
        unique_code += '-'

        if is_instrument:
            unique_code += 'I' + str(released_material.id)
        else:
            unique_code += 'M' + str(released_material.id)
        released_material.unique_code = unique_code
        released_material.save()
        try:
            indexs = list(range(1, ReleasedMaterialItem.objects.filter(released_material=released_material).count() + 1))
            context = {
                'object_name': ConstructionObject.objects.get(contract=contract).name,
                'object_address': ConstructionObject.objects.get(contract=contract).address,
                'number_contract': contract.number_contract,
                'date_contract': contract.date_contract,
                'date_doc': datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
                'number_doc': released_material.unique_code,
                'role': request.user.get_role_display(),
                'name': request.user.first_name + ' ' + request.user.last_name,
                'contract_contractor': contract.contractor,
                'materials': ReleasedMaterialItem.objects.filter(released_material=released_material),
                'indexs': indexs,
                'contract_name': contract.name,

            }

            tpl = DocxTemplate(os.path.join(BASE_DIR, 'mediafiles/nakladnaya.docx'))
            tpl.render(context)
            tpl.save('mediafiles/waybill/nakladnaya-' + contract.name + str(released_material.id) + '.docx')

        except:
            return HttpResponse('Ошибка')
    return redirect(request.META.get('HTTP_REFERER'))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ReleasedMaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/relesed_materials_list.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(contract__slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        self.extra_context = {
            'construction_object': construction_object,
            'relesed_materials': ReleasedMaterial.objects.filter(contract__slug=self.kwargs['slug'],items__material__instrument_code=None).order_by('id').distinct()
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ReturnReleaseMaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/return_release_mat.html'

    def get(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        construction_object = released_material.contract.construction_object
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        self.extra_context = {
            'construction_object': construction_object,
            'released_material': released_material,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        contract = Contract.objects.get(realeas_material__id=int(self.kwargs['id']))
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        released_material.is_done = True
        released_material.save()
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            released_material_item = ReleasedMaterialItem.objects.get(material=material, released_material=released_material)
            return_material = request.POST['return_mat' + str(i)]
            material.release_count = material.release_count - int(return_material)
            material.save()
            released_material_item.return_count = return_material
            released_material_item.save()

        try:
            indexs = list(range(1, ReleasedMaterialItem.objects.filter(released_material=released_material).count() + 1))
            context = {
                'object_name': ConstructionObject.objects.get(contract=contract).name,
                'object_address': ConstructionObject.objects.get(contract=contract).address,
                'number_contract': contract.number_contract,
                'date_contract': contract.date_contract,
                'date_doc': datetime.datetime.now().strftime("%d-%m-%Y %H:%M"),
                'number_doc': released_material.unique_code,
                'role': request.user.get_role_display(),
                'name': request.user.first_name + ' ' + request.user.last_name,
                'contract_contractor': contract.contractor,
                'materials': ReleasedMaterialItem.objects.filter(released_material=released_material),
                'indexs': indexs,
                'contract_name': contract.name,
            }
            tpl = DocxTemplate(os.path.join(BASE_DIR, 'mediafiles/nakladnaya_final.docx'))
            tpl.render(context)
            tpl.save('mediafiles/waybill/nakladnaya_final-' + contract.name + str(released_material.id) + '.docx')
        except:
            return HttpResponse('Ошибка')
        if released_material.items.all()[0].material.instrument_code == None:
            return redirect('/contract/' + contract.slug + '/relesed_materials')
        else:
            return redirect('/construction_objects/' + contract.construction_object.slug + '/released_instruments')


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class DetailReleaseMaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/detial_released_mat.html'

    def get(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        construction_object = released_material.contract.construction_object
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        self.extra_context = {
            'construction_object': construction_object,
            'released_material': released_material,
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AddReleaseWaybillView(generic.TemplateView):
    template_name = 'materials/store_mateials/waybill/add_release_waybill.html'

    def get(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        construction_object = released_material.contract.construction_object
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        self.extra_context = {
            'construction_object': construction_object,
            'released_material': released_material,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        released_waybill = request.FILES['waybill']
        released_material.release_waybill = released_waybill
        released_material.save()
        return redirect('/detail/relesed_materials/' + str(self.kwargs['id']))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AddFinalWaybillView(generic.TemplateView):
    template_name = 'materials/store_mateials/waybill/add_final_waybill.html'

    def get(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        construction_object = released_material.contract.construction_object
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        self.extra_context = {
            'construction_object': construction_object,
            'released_material': released_material,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        released_material = ReleasedMaterial.objects.get(id=int(self.kwargs['id']))
        final_waybill = request.FILES['waybill']
        released_material.final_waybill = final_waybill
        released_material.save()
        return redirect('/detail/relesed_materials/' + str(self.kwargs['id']))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InstrumentMateriralView(generic.TemplateView):
    template_name = 'materials/store_mateials/instruments/instruments.html'

    def get(self, request, *args, **kwargs):
        contstruction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or contstruction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        materials = Material.objects.filter(is_delivery=True, invoice__is_done=True,invoice__request_for_material__contract__construction_object=contstruction_object).filter(
            ~Q(instrument_code=None))

        self.extra_context = {
            'construction_object': contstruction_object,
            'materials': materials,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        materials = request.POST.getlist('materials')
        materials = Material.objects.filter(id__in=materials)
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        context = {
            'materials': materials,
            'construction_object': construction_object,
        }
        return render(request, template_name='materials/store_mateials/release_materials.html',context=context)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ReleasedInstruments(generic.TemplateView):
    template_name = 'materials/store_mateials/instruments/released_instruments.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        self.extra_context = {
            'construction_object': construction_object,
            'relesed_materials': ReleasedMaterial.objects.filter(contract__construction_object=construction_object).filter(~Q(items__material__instrument_code=None))
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class GeneralBaseView(generic.TemplateView):
    template_name = 'materials/store_mateials/general_base.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        materials = Material.objects.filter(is_delivery=True, invoice__is_done=True, invoice__request_for_material__contract__construction_object=construction_object)
        self.extra_context = {
            'construction_object': construction_object,
            'materials': materials,
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class RemainderMaterialsView(generic.TemplateView):
    template_name = 'materials/store_mateials/remainder_materials.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        materials = Material.objects.filter(is_delivery=True, invoice__is_done=True, invoice__request_for_material__contract__construction_object=construction_object)
        self.extra_context = {
            'construction_object': construction_object,
            'materials': materials,
        }
        return super().get(request, *args, **kwargs)
