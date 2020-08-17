from django.shortcuts import render, redirect, HttpResponse
from django.views import generic
from .models import Object, Contract, Material, RequestForMaterial, InvoiceForPayment
from transliterate import slugify
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from openpyxl import load_workbook
import requests

general_bot_token = '1270115367:AAGCRLBP1iSZhpTniwVYQ9p9fqLysY668ew'
status_dict = {
    'в работе': 1,
    'невыполнено': 2,
    'проверка': 3,
    'выполнено': 4
}


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class HomeView(generic.ListView):
    template_name = 'index.html'
    queryset = Object.objects.all()


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ObjectDetailView(generic.TemplateView):
    template_name = 'appbase/object_detail.html'

    # model = Object

    def get(self, request, *args, **kwargs):
        object_slug = self.kwargs['slug']
        pk = self.kwargs['pk']
        status = ''
        if pk == 1:
            contracts = Contract.objects.filter(contstruct_object__slug=object_slug, status='в работе')
            status = 'в работе'
        elif pk == 2:
            contracts = Contract.objects.filter(contstruct_object__slug=object_slug, status='невыполнено')
            status = 'невыполнено'
        elif pk == 3:
            contracts = Contract.objects.filter(contstruct_object__slug=object_slug, status='проверка')
            status = 'проверка'
        elif pk == 4:
            contracts = Contract.objects.filter(contstruct_object__slug=object_slug, status='выполнено')
            status = 'выполнено'

        self.extra_context = {
            'contracts': contracts,
            'status': status,
            'object': Object.objects.get(slug=object_slug)
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ContractDetailView(generic.TemplateView):
    template_name = 'appbase/contract/detail.html'

    # model = Material
    # queryset = Material.objects.all()

    def get(self, request, *args, **kwargs):
        contract_slug = self.kwargs['slug']
        contract = Contract.objects.get(slug=contract_slug)

        self.extra_context = {
            'contract': contract,
            'requests': RequestForMaterial.objects.filter(contract=contract),
            'object': Object.objects.get(slug=contract.contstruct_object.slug),

        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ContractAddView(generic.TemplateView):
    template_name = 'appbase/contract/add.html'

    def get(self, request, *args, **kwargs):
        print('asd')
        self.extra_context = {
            'object': Object.objects.get(slug=self.kwargs['slug']),
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        object_id = request.POST['object']
        name = request.POST['name']
        contractor = request.POST['contractor']
        contract = request.FILES['contract']
        number_contract = request.POST['number_contract']
        status = request.POST['status']
        if slugify(name) == None:
            slug = name
        else:
            slug = slugify(name)

        Contract.objects.create(contstruct_object_id=object_id, name=name, slug=slug, contract=contract,
                                contractor=contractor, number_contract=number_contract, status=status)

        return redirect('/objects/' + Object.objects.get(id=object_id).slug + '/' + str(status_dict[status]))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ContractEditView(generic.TemplateView):
    template_name = 'appbase/contract/edit.html'

    def get(self, request, *args, **kwargs):

        contract_slug = self.kwargs['slug']

        contract = Contract.objects.get(slug=contract_slug)
        self.extra_context = {
            'object': Object.objects.get(slug=contract.contstruct_object.slug),
            'contract': contract
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        contract_slug = self.kwargs['slug']
        contract = Contract.objects.get(slug=contract_slug)
        object_id = request.POST['object']
        name = request.POST['name']
        if request.FILES:
            contract_file = request.FILES['contract']
        else:
            contract_file = contract.contract
        contractor = request.POST['contractor']

        number_contract = request.POST['number_contract']
        status = request.POST['status']
        slug = slugify(name)

        contract.name = name
        contract.contract = contractor
        contract.contract = contract_file
        contract.number_contract = number_contract
        contract.status = status
        contract.save()

        return redirect('/objects/' + Object.objects.get(id=object_id).slug + '/' + str(status_dict[status]))


@login_required(login_url='/accounts/login/')
def contract_delete(request):
    contract = Contract.objects.get(id=int(request.POST['contract']))
    red = contract.contstruct_object.slug + '/' + str(status_dict[contract.status])
    contract.delete()

    return redirect(red)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class RequestAddView(generic.TemplateView):
    template_name = 'appbase/contract/request/add.html'

    def get(self, request, *args, **kwargs):
        contract = Contract.objects.get(slug=self.kwargs['slug'])
        self.extra_context = {
            'contract': contract,
            'object': Object.objects.get(contract=contract)
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        contract = Contract.objects.get(id=request.POST['contract'])
        name = request.POST['name']
        file = request.FILES['request']
        RequestForMaterial.objects.create(contract=contract, name=name, file=file)
        return redirect('/contract/detail/' + contract.slug)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class RequestEditView(generic.TemplateView):
    template_name = 'appbase/contract/request/edit.html'

    def get(self, request, *args, **kwargs):
        request_mat = RequestForMaterial.objects.get(id=self.kwargs['id'])
        self.extra_context = {
            'request_mat': request_mat,
            'object': Object.objects.get(contract__request=request_mat)
        }

        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        request_mat = RequestForMaterial.objects.get(id=request.POST['id'])

        # contract_slug = self.kwargs['contract_slug']
        # contract = Contract.objects.get(slug=contract_slug)
        # object_id = request.POST['object']
        name = request.POST['name']
        if request.FILES:
            request_mat_file = request.FILES['request']
        else:
            request_mat_file = request_mat.file

        request_mat.name = name
        request_mat.file = request_mat_file
        request_mat.save()

        return redirect('/contract/detail/' + request_mat.contract.slug)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class RequestDetailView(generic.TemplateView):
    template_name = 'appbase/contract/request/detail.html'

    def get(self, request, *args, **kwargs):
        request_mat = RequestForMaterial.objects.get(id=self.kwargs['id'])

        self.extra_context = {
            'request_mat': request_mat,
            'object': Object.objects.get(slug=request_mat.contract.contstruct_object.slug),
        }
        return super().get(request, *args, **kwargs)


@login_required(login_url='/accounts/login/')
def request_delete(request):
    request_mat = RequestForMaterial.objects.get(id=request.POST['id'])
    red = '/contract/detail/' + request_mat.contract.slug
    request_mat.delete()

    return redirect(red)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InvoiceAddView(generic.TemplateView):
    template_name = 'appbase/contract/request/invoice/add.html'

    def get(self, request, *args, **kwargs):
        request_mat = RequestForMaterial.objects.get(id=self.kwargs['id'])
        self.extra_context = {
            'request_mat': request_mat,
            'object': Object.objects.get(contract__request=request_mat)
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        request_mat = RequestForMaterial.objects.get(id=int(request.POST['id']))
        file = request.FILES['invoice']
        bin = request.POST['bin']
        name_company = request.POST['name']
        comment = request.POST['comment']

        InvoiceForPayment.objects.create(request_mat=request_mat, file=file, bin=bin, name_company=name_company,
                                         comment=comment)
        return redirect('/request/detail/' + str(request_mat.id))


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InvoiceEditView(generic.TemplateView):
    template_name = 'appbase/contract/request/invoice/edit.html'

    def get(self, request, *args, **kwargs):
        invoice = InvoiceForPayment.objects.get(id=int(self.kwargs['id']))

        self.extra_context = {
            'invoice': invoice,
            'object': Object.objects.get(contract__request__invoice=invoice)
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        invoice = InvoiceForPayment.objects.get(id=int(request.POST['id']))
        file = request.FILES['invoice']
        invoice.file = file
        invoice.save()
        return redirect('/request/detail/' + str(invoice.request_mat.id))


@login_required(login_url='/accounts/login/')
def invoice_delete(request):
    invoice = InvoiceForPayment.objects.get(id=request.POST['id'])
    red = '/request/detail/' + str(invoice.request_mat.id)
    invoice.delete()

    return redirect(red)


import datetime


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InvoiceForPaymentView(generic.TemplateView):
    template_name = 'appbase/invoices.html'

    def get(self, request, *args, **kwargs):
        con_object = Object.objects.get(slug=self.kwargs['slug'])
        invoices = InvoiceForPayment.objects.filter(status='да', request_mat__contract__contstruct_object=con_object)
        print(datetime.datetime.now())
        self.extra_context = {
            'object': con_object,
            'invoices': invoices,
            'current_date': datetime.datetime.now()

        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        invoice = InvoiceForPayment.objects.get(id=int(request.POST['invoice_id']))

        if request.POST['submit'] == 'paid':
            invoice.is_paid = True
            invoice.reset_date = datetime.datetime.now() + datetime.timedelta(hours=1)
            invoice.save()
        elif request.POST['submit'] == 'looked':
            invoice.is_looked = True
            invoice.save()
        elif request.POST['submit'] == 'cancel':
            invoice.is_looked = False
            invoice.is_paid = False
            invoice.save()
        return redirect('/invoice_for_payment/' + invoice.request_mat.contract.contstruct_object.slug)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AddMaterialView(generic.TemplateView):
    template_name = 'appbase/contract/request/invoice/add_material.html'

    def get(self, request, *args, **kwargs):
        self.extra_context = {
            'object': Object.objects.get(contract__request__invoice__id=self.kwargs['id']),
            # 'contract': Contract.objects.get(slug=self.kwargs['slug']),
            # 'request_mat': RequestForMaterial.objects.get(id=self.kwargs['id'])
            'invoice': InvoiceForPayment.objects.get(id=self.kwargs['id'])
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # contract = Contract.objects.get(id=int(request.POST['contract']))
        invoice = InvoiceForPayment.objects.get(id=int(request.POST['id']))

        file = request.FILES['file']
        wb = load_workbook(file)
        sheet_name = wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        data = list(sheet.values)
        for item in data:
            name = item[0]
            quantuty = item[1]
            units = item[2]
            price = item[3]
            sum_price = item[4]
            Material.objects.create(invoice=invoice, name=name, quantity=quantuty, units=units, price=price,
                                    sum_price=sum_price)

        return redirect('/request/detail/' + str(invoice.request_mat.id))
        # return super().get(request, *args, **kwargs)


class PaidMaterailsView(generic.TemplateView):
    template_name = 'appbase/material/invoices.html'

    def get(self, request, *args, **kwargs):
        # materails = Material.objects.filter(request_mat__contract__contstruct_object__slug=self.kwargs['slug'],
        #                                     is_delivery=False)
        invoices = InvoiceForPayment.objects.filter(request_mat__contract__contstruct_object__slug=self.kwargs['slug'])

        self.extra_context = {
            'object': Object.objects.get(slug=self.kwargs['slug']),
            'invoices': invoices,

            # 'materials': materails
        }
        return super().get(request, *args, **kwargs)



class InvoicePaidMaterialsView(generic.TemplateView):
    template_name = 'appbase/material/paid_materials.html'

    def get(self, request, *args, **kwargs):
        materials = Material.objects.filter(invoice__id=int(self.kwargs['id']))
        self.extra_context = {
            'object': Object.objects.get(contract__request__invoice__id=self.kwargs['id']),
            'materials': materials,
            'invoice': InvoiceForPayment.objects.get(id=self.kwargs['id']),
        }
        return super().get(request, *args, **kwargs)


    def post(self, request, *args, **kwargs):
        materials = request.POST.getlist('materials')
        if request.POST['submit'] == 'delivered':
            for material_id in materials:
                material = Material.objects.get(id=material_id)
                material.is_delivery = True
                material.status = 'ок'
                # material.save()

        elif request.POST['submit'] == 'marriage':
            # status = request.POST['marriage']
            materials = Material.objects.filter(id__in=materials)
            context = {
                'materials' : materials,
                'object': Object.objects.get(contract__request__invoice__id=self.kwargs['id']),
            }
            return render(request, template_name='appbase/material/marriage_materials.html', context=context)
            # for material_id in materials:
            #     material = Material.objects.get(id=material_id)
            #     material.status = status
            #     material.save()


            # requests.get("https://api.telegram.org/bot%s/sendMessage" % general_bot_token,
            #              params={'chat_id': '-1001342160485', 'text': message})
        return redirect('/objects/invoice/' + str(self.kwargs['id']) + '/materials')

class MaterialsView(generic.TemplateView):
    template_name = 'appbase/material/materials.html'

    def get(self, request, *args, **kwargs):
        self.extra_context = {
            'object': Object.objects.get(slug=self.kwargs['slug']),
            'materials': Material.objects.filter(request_mat__contract__contstruct_object__slug=self.kwargs['slug'],
                                                 is_delivery=True)
        }
        return super().get(request, *args, **kwargs)


import telebot
from telebot import types

token = '1318683651:AAH_fhHdb-PGt9kGhSqEOrVXvak3-jFRljk'

channel_id = '-1001342160485'

bot = telebot.TeleBot(token)


def send_telegram(request):
    keyboard = types.InlineKeyboardMarkup()  # наша клавиатура
    key_yes = types.InlineKeyboardButton(text='Да', callback_data='yes')  # кнопка «Да»
    keyboard.add(key_yes)  # добавляем кнопку в клавиатуру
    key_no = types.InlineKeyboardButton(text='Нет', callback_data='no')
    keyboard.add(key_no)
    key_then = types.InlineKeyboardButton(text='Потом', callback_data='then')
    keyboard.add(key_then)

    invoice = InvoiceForPayment.objects.get(id=int(request.POST['id']))
    request_mat_file = invoice.request_mat.file
    invoice_file = invoice.file
    msg = '\n\n\n'
    msg += '🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔\n'
    msg += 'Новый счет!\n'
    bot.send_message('438797738', text=msg)
    msg = 'Объект: ' + str(invoice.request_mat.contract.contstruct_object) + '\n'
    msg += 'Работа: ' + str(invoice.request_mat.contract) + '\n'
    msg += 'Текущий статус: ' + str(invoice.status) + '\n'
    bot.send_message('438797738', text=msg)

    bot.send_document('438797738', request_mat_file)
    bot.send_document('438797738', invoice_file)
    msg = 'ID: ' + str(invoice.id) + '\n'
    bot.send_message('438797738', text=msg, reply_markup=keyboard)

    return redirect('/request/detail/' + str(invoice.request_mat.id))


def api_telegram_response(request):
    message = 'Счет на оплату!' + '\n'
    bot = telebot.TeleBot(general_bot_token)

    if request.GET['token'] == '123':
        response = request.GET['response']
        invoice_id = int(request.GET['invoice_id'])
        invoice = InvoiceForPayment.objects.get(id=invoice_id)
        status = '-'
        if response == 'yes':
            status = 'да'
        elif response == 'no':
            status = 'нет'
        elif response == 'then':
            status = 'потом'
        print(status)
        invoice.status = status
        invoice.save()
        message += 'Ответ: ' + status
        bot.send_message(channel_id, text=message)
        bot.send_document(channel_id, invoice.request_mat.file)
        bot.send_document(channel_id, invoice.file)

    # requests.get("https://api.telegram.org/bot%s/sendMessage" % general_bot_token,
    #              params={'chat_id': '-1001302242759', 'text': message})
    return HttpResponse('ok')
