from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import generic
from construction_objects_app.models import ConstructionObject
from contracts_app.models import InvoiceForPayment
from .models import Material
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from openpyxl import load_workbook
import requests
from django.db.models import F, Q

import telebot

from .forms import MaterialForm
from django.urls import reverse
from contracts_app.views import check_user

GENERAL_BOT_TOKEN = '1270115367:AAGCRLBP1iSZhpTniwVYQ9p9fqLysY668ew'
channel_id = '-1001342160485'


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AddMaterialsExcelView(generic.TemplateView):
    template_name = 'paid_material_app/add_materials_excel.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        invoice = InvoiceForPayment.objects.get(id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'manager' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        self.extra_context = {
            'construction_object': construction_object,
            'invoice': invoice,
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):

        invoice = InvoiceForPayment.objects.get(id=int(request.POST['id']))
        doc_file = request.FILES['doc_file']
        wb = load_workbook(doc_file)
        sheet_name = wb.sheetnames[0]
        sheet = wb.get_sheet_by_name(sheet_name)
        data = list(sheet.values)
        for item in data:
            if item[0] is None:
                continue
            name = item[0]
            quantuty = item[1]
            units = item[2]
            price = item[3]
            sum_price = item[4]
            instriment_code = None
            if len(item) > 5:
                instriment_code = item[5]
            if (isinstance(quantuty, int)) == False:
                quantuty = int(quantuty)
            if (isinstance(price, int)) == False:
                price = int(price)
            sum_price = quantuty * price
            material = Material.objects.create(invoice=invoice, name=name, quantity=quantuty, units=units, price=price,
                                               sum_price=int(sum_price))

            if instriment_code == 1:
                material.is_instrument = True
            material.save()
        return redirect('/request/' + str(invoice.request_for_material.id) + '/detail/')


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class MaterialCreateView(generic.CreateView):
    model = Material
    form_class = MaterialForm

    def get_success_url(self, **kwargs):
        return reverse('invoice_for_payment_detail', kwargs={'pk': self.kwargs['pk']})

    def form_valid(self, form):
        invoice = InvoiceForPayment.objects.get(id=self.kwargs['pk'])
        form.instance.invoice = invoice
        return super(MaterialCreateView, self).form_valid(form)

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['pk'])
        if check_user(request.user, ['accountant', 'manager'], construction_object) == 404:
            return render(request, '404.html')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['pk'])
        invoice = InvoiceForPayment.objects.get(id=self.kwargs['pk'])
        context['construction_object'] = construction_object
        context['invoice'] = invoice
        return context


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class MaterialUpdateView(generic.UpdateView):
    model = Material
    form_class = MaterialForm

    def get_success_url(self, **kwargs):
        return reverse('invoice_for_payment_detail',
                       kwargs={'pk': InvoiceForPayment.objects.get(material__id=self.kwargs['pk']).id})

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__material__id=self.kwargs['pk'])
        if check_user(request.user, ['accountant', 'manager'], construction_object) == 404:
            return render(request, '404.html')
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__material__id=self.kwargs['pk'])
        invoice = InvoiceForPayment.objects.get(material__id=self.kwargs['pk'])
        context['construction_object'] = construction_object
        context['invoice'] = invoice
        return context


@login_required(login_url='/accounts/login/')
def material_delete(request):
    material = Material.objects.get(id=int(request.POST['material_id']))
    invoice = material.invoice
    construction_object = invoice.request_for_material.contract.construction_object
    if request.user.role == 'accountant' or request.user.role == 'manager' or construction_object not in list(
            request.user.construction_objects.all()):
        return render(request, template_name='404.html')
    red = '/invoice/' + str(invoice.id) + '/detail/'
    material.delete()
    return redirect(red)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class PaidMaterailsView(generic.TemplateView):
    template_name = 'paid_material_app/invoices.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')

        invoices = InvoiceForPayment.objects.filter(request_for_material__contract__construction_object=construction_object).filter(~Q(name_company=construction_object.name))

        self.extra_context = {
            'construction_object': construction_object,
            'invoices': invoices,
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class InvoicePaidMaterialsView(generic.TemplateView):
    template_name = 'paid_material_app/paid_materials.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        materials = Material.objects.filter(invoice__id=int(self.kwargs['id'])).filter(~Q(ok=F('quantity')))
        # materials = Material.objects.filter(invoice__id=int(self.kwargs['id']), is_delivery=False)

        invoice = InvoiceForPayment.objects.get(id=self.kwargs['id'])

        if list(materials) == [] and list(invoice.material.all()) != []:
            invoice.is_done = True
            invoice.save()

        self.extra_context = {
            'construction_object': construction_object,
            'materials': materials,
            'invoice': invoice
        }
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment__id=self.kwargs['id'])
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        materials = request.POST.getlist('materials')
        materials = Material.objects.filter(id__in=materials)
        context = {
            'materials': materials,
            'construction_object': construction_object,
        }
        if request.POST['submit'] == 'delivered':

            for material in materials:
                material.is_delivery = True
                material.status = 'ок'
                material.ok = material.quantity
                material.remainder_count = material.quantity
                material.marriage, material.shortage, material.inconsistency = 0, 0, 0
                material.save()
        elif request.POST['submit'] == 'marriage':
            return render(request, template_name='paid_material_app/marriage_materials.html',
                          context=context)

        elif request.POST['submit'] == 'return':
            return render(request, template_name='paid_material_app/return_materials.html',
                          context=context)
        return redirect('/construction_objects/invoice/' + str(self.kwargs['id']) + '/materials')


def marriage_materials(request):
    if request.POST:
        message = '🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔\n'
        message += 'Бракованные товары\n'
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            material.ok = request.POST['ok' + str(i)]
            material.marriage = request.POST['brak' + str(i)]
            material.shortage = request.POST['neso' + str(i)]
            material.inconsistency = request.POST['nexv' + str(i)]
            material.status = 'брак'
            if int(material.marriage) + int(material.ok) + int(material.shortage) + int(
                    material.inconsistency) != material.quantity:
                return HttpResponse('Ошибка!')
            if int(material.ok) > 1:
                material.is_delivery = True
            material.save()

            message += str(i) + '. ' + material.name + '\n' + 'Ок: ' + str(
                material.ok) + '\nБрак: ' + material.marriage + '\nНехватка: ' + material.inconsistency + '\nНесоответсвие: ' + material.shortage + '\n\n'
            invoice = material.invoice
        if request.POST['comment']:
            message += 'Примечение: ' + request.POST['comment'] + '\n'
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment=invoice)
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        message += 'БИН: ' + invoice.bin + '\n'
        message += 'Название компании: ' + invoice.name_company + '\n'
        requests.get("https://api.telegram.org/bot%s/sendMessage" % GENERAL_BOT_TOKEN,
                     params={'chat_id': '-1001342160485', 'text': message})
        bot = telebot.TeleBot(GENERAL_BOT_TOKEN)
        bot.send_document(channel_id, invoice.doc_file)
    return redirect(request.META.get('HTTP_REFERER'))


def return_materials(request):
    if request.POST:
        message = '🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔🔔\n'
        message += 'Возврат товаров\n'
        for i in range(1, int(request.POST['count']) + 1):
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            return_materials_count = int(request.POST['return' + str(i)])
            invoice = material.invoice
            if return_materials_count > material.quantity:
                return HttpResponse('Ошибка!')
            elif return_materials_count == material.quantity:
                material.delete()
            else:
                material.quantity = material.quantity - return_materials_count
                material.marriage, material.shortage, material.inconsistency = 0, 0, 0
                material.ok = material.quantity
                material.remainder_count = material.quantity
                material.status = 'ок'
                material.save()
            message += str(i) + '. ' + material.name + '\n' + 'Количество: ' + str(
                material.quantity) + '\nКоличество возвратных товаров: ' + str(return_materials_count) + '\n\n'
        if request.POST['comment']:
            message += 'Примечение: ' + request.POST['comment']
        construction_object = ConstructionObject.objects.get(
            contract__request_for_material__invoice_for_payment=invoice)
        if request.user.role == 'accountant' or request.user.role == 'purchaser' or construction_object not in list(
                request.user.construction_objects.all()):
            return render(request, template_name='404.html')
        message += 'БИН: ' + str(invoice.bin) + '\n'
        message += 'Название компании: ' + invoice.name_company + '\n'
        requests.get("https://api.telegram.org/bot%s/sendMessage" % GENERAL_BOT_TOKEN,
                     params={'chat_id': '-1001342160485', 'text': message})
        bot = telebot.TeleBot(GENERAL_BOT_TOKEN)
        bot.send_document(channel_id, invoice.doc_file)
    return redirect(request.META.get('HTTP_REFERER'))
