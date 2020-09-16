from django.shortcuts import redirect
from django.http import HttpResponse, FileResponse
from django.views import generic
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from construction_objects_app.models import ConstructionObject, Contract, RequestForMaterial
from material_app.models import Material, ReleasedMaterial
from .filters import MaterialFilter, ReleasedMaterialFilter
from django.db.models import Sum
import datetime
import pytz
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Color
from docxtpl import DocxTemplate
import os
import io
from foldable_base.settings import BASE_DIR


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class AnalyticsView(generic.ListView):
    template_name = 'analytics/analytics.html'
    model = Material

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset().filter(is_delivery=True, invoice__is_done=True,
                                              invoice__request_for_material__contract__construction_object__slug=
                                              self.kwargs['slug'])
        if list(queryset) == []:
            return context
        material_filter = MaterialFilter(self.request.GET, queryset=queryset)
        context['total_sum_price'] = round((material_filter.qs.aggregate(Sum('sum_price'))['sum_price__sum']), 2)
        context['filter'] = material_filter

        return context

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        self.extra_context = {
            'construction_object': construction_object,
        }
        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class TotalStats(generic.TemplateView):
    template_name = 'analytics/total_stats.html'

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        contracts = Contract.objects.filter(construction_object=construction_object)
        request_mats = RequestForMaterial.objects.filter(contract__construction_object=construction_object)
        materials = Material.objects.filter(
            invoice__request_for_material__contract__construction_object=construction_object, is_delivery=True,
            invoice__is_done=True)

        total_sum_price = round(materials.aggregate(Sum('sum_price'))['sum_price__sum'])

        is_cash_total_sum_price = round(
            materials.filter(invoice__is_cash=True).aggregate(Sum('sum_price'))['sum_price__sum'])

        start_date_default = datetime.datetime(2020, 8, 1, tzinfo=pytz.UTC)
        end_date_default = datetime.datetime.today().date()
        in_work = contracts.filter(status='1')
        finished = contracts.filter(status='2')
        not_finished = contracts.filter(status='3')
        check = contracts.filter(status='4')
        is_done_request_mat = request_mats.filter(is_done=True)
        if 'start_date' in request.GET:
            start_date_default = request.GET['start_date'].replace('-', '/')
            end_date_default = request.GET['end_date'].replace('-', '/')
            start_date_default = datetime.datetime.strptime(start_date_default, '%Y/%m/%d')
            end_date_default = datetime.datetime.strptime(end_date_default, '%Y/%m/%d')
            contracts = contracts.filter(created_at__gte=start_date_default,
                                         created_at__lte=end_date_default)
            request_mats = request_mats.filter(created_at__gte=start_date_default,
                                               created_at__lte=end_date_default)
            materials = materials.filter(created_at__gte=start_date_default,
                                         created_at__lte=end_date_default + datetime.timedelta(days=1))

            in_work = contracts.filter(status='1')
            finished = contracts.filter(status='2')
            not_finished = contracts.filter(status='3')
            check = contracts.filter(status='4')
            is_done_request_mat = request_mats.filter(is_done=True)
            total_sum_price = materials.aggregate(Sum('sum_price'))['sum_price__sum']

        self.extra_context = {
            'construction_object': construction_object,
            'contracts': contracts,
            'request_mats': request_mats,
            'materials': materials,
            'total_sum_price': total_sum_price,
            'start_date_default': start_date_default,
            'end_date_default': end_date_default,
            'in_work': in_work,
            'finished': finished,
            'not_finished': not_finished,
            'check': check,
            'is_done_request_mat': is_done_request_mat,
            'is_cash_total_sum_price': is_cash_total_sum_price

        }

        return super().get(request, *args, **kwargs)


@method_decorator(login_required(login_url='/accounts/login/'), name='dispatch')
class ReleasedMaterialsStats(generic.ListView):
    template_name = 'analytics/release_mat_stats.html'
    model = ReleasedMaterial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        queryset = queryset.filter(contract__construction_object=construction_object).order_by('-release_date')
        if list(queryset) == []:
            return context
        self.queryset = queryset
        release_material_filter = ReleasedMaterialFilter(self.request.GET, queryset=queryset)
        context['filter'] = release_material_filter

        return context

    def get(self, request, *args, **kwargs):
        construction_object = ConstructionObject.objects.get(slug=self.kwargs['slug'])
        self.extra_context = {
            'construction_object': construction_object,
        }
        return super().get(request, *args, **kwargs)


def export_analytics(request, slug):
    if request.POST:
        construction_object = ConstructionObject.objects.get(id=int(request.POST['construction_object_id']))
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']

        if start_date == '':
            start_date = str(construction_object.created_at.date())
        if end_date == '':
            end_date = str(datetime.datetime.today().date())
        start_date = str(start_date).split('-')
        start_date = start_date[2] + '.' + start_date[1] + '.' + start_date[0]
        end_date = str(end_date).split('-')
        end_date = end_date[2] + '.' + end_date[1] + '.' + end_date[0]
        wb = load_workbook('mediafiles/analytics.xlsx')
        ws = wb.active

        count_materials = int(request.POST['count_materials'])

        ws['B2'] = 'Отчет материальных ценностей по объекту "' + construction_object.name + '"'
        ws['C4'] = start_date
        ws['C5'] = end_date
        for row in ws["A9:M" + str(count_materials + 8)]:
            for cell in row:
                cell.style = 'Output'
        for i in range(1, count_materials + 1):
            request_for_material_done = 'Нет'
            invoice_is_cash = 'Нет'
            material = Material.objects.get(id=int(request.POST['material' + str(i)]))
            instrument_code = material.instrument_code
            if material.invoice.request_for_material.is_done:
                request_for_material_done = 'Да'
            if material.invoice.is_cash:
                invoice_is_cash = 'Да'
            if material.instrument_code is None:
                instrument_code = 'Нет'

            ws['A' + str(i + 8)] = i
            ws['B' + str(i + 8)] = material.name
            ws['C' + str(i + 8)] = material.invoice.request_for_material.contract.name
            ws['D' + str(i + 8)] = material.invoice.request_for_material.contract.get_status_display()
            ws['E' + str(i + 8)] = request_for_material_done
            ws['F' + str(i + 8)] = material.quantity
            ws['G' + str(i + 8)] = material.release_count
            ws['H' + str(i + 8)] = material.quantity - material.release_count
            ws['I' + str(i + 8)] = material.units
            ws['K' + str(i + 8)] = instrument_code
            ws['K' + str(i + 8)] = invoice_is_cash
            ws['L' + str(i + 8)] = material.price
            ws['M' + str(i + 8)] = material.sum_price
        ws['A' + str(count_materials + 9)].font = Font(bold=True)
        ws['M' + str(count_materials + 9)].font = Font(bold=True)
        ws['A' + str(count_materials + 9)] = 'Итого: ' + str(count_materials)
        ws['M' + str(count_materials + 9)] = request.POST['total_sum_price']

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=analytics.xlsx'
        return response


def export_total_stats(request):
    if request.POST:
        context = {
            'construction_object_name': request.POST['construction_object_name'],
            'start_date': request.POST['start_date'],
            'end_date': request.POST['start_date'],
            'contracts_count': request.POST['contracts_count'],
            'in_work_count': request.POST['in_work_count'],
            'finished_count': request.POST['finished_count'],
            'not_finished_count': int(request.POST['not_finished_count']),
            'request_for_materials_count': request.POST['request_for_materials_count'],
            'request_for_materials_is_done_count': request.POST['request_for_materials_is_done_count'],
            'materials_count': request.POST['materials_count'],
            'total_sum_price': request.POST['total_sum_price'],
            'is_cash_total_sum_price': request.POST['is_cash_total_sum_price'],

        }
        byte_io = io.BytesIO()
        tpl = DocxTemplate(os.path.join(BASE_DIR, 'mediafiles/total_stats.docx'))
        tpl.render(context)
        tpl.save(byte_io)
        byte_io.seek(0)
        return FileResponse(byte_io, as_attachment=True, filename='total_stats.docx')


def export_release_mat_stats(request):
    if request.POST:
        construction_object = ConstructionObject.objects.get(id=int(request.POST['construction_object_id']))
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']

        if start_date == '':
            start_date = str(construction_object.created_at.date())
        if end_date == '':
            end_date = str(datetime.datetime.today().date())
        start_date = str(start_date).split('-')
        start_date = start_date[2] + '.' + start_date[1] + '.' + start_date[0]
        end_date = str(end_date).split('-')
        end_date = end_date[2] + '.' + end_date[1] + '.' + end_date[0]
        wb = load_workbook('mediafiles/release_mat_stats.xlsx')
        ws = wb.active
        ws['C2'] = 'Отчет движения материальных ценностей по объекту "' + construction_object.name + '"'
        ws['D4'] = start_date
        ws['D5'] = end_date
        count_release_material = request.POST['count_release_material']
        a = 9
        for i in range(1, int(count_release_material) + 1):
            release_material_id = request.POST['release_material' + str(i)]
            release_material = ReleasedMaterial.objects.get(id=release_material_id)
            count_materials = release_material.items.count()
            ws['B' + str(a)] = release_material.release_date.astimezone(pytz.timezone('Asia/Almaty')).strftime(
                "%d.%m.%Y %H:%M ")
            ws['A' + str(a)] = release_material.unique_code

            for material in release_material.items.all():
                instrument_code = material.material.instrument_code
                if material.material.instrument_code is None:
                    instrument_code = 'Нет'
                ws['C' + str(a)] = material.material.name
                ws['D' + str(a)] = material.material.invoice.request_for_material.contract.name
                ws['E' + str(a)] = material.material.quantity
                ws['F' + str(a)] = material.release_count
                ws['G' + str(a)] = material.return_count
                ws['H' + str(a)] = material.material.quantity - material.release_count + material.return_count
                ws['I' + str(a)] = material.material.units
                ws['J' + str(a)] = instrument_code
                a += 1
        for row in ws["A9:J" + str(a-1)]:
            for cell in row:
                cell.style = 'Output'

        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=hod_dvizhenij.xlsx'
        return response
        # return redirect('/construction_objects/nurly-tau/released_material_stats')